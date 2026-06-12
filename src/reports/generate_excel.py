# src/reports/generate_excel.py
import pandas as pd
import sys
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from config.config import DATA_PROCESSED, DATA_EXPORTS
from src.analysis.kpis import (
    load_dataset,
    kpis_generales,
    comparativa_anual,
    ventas_por_mes,
    top_productos,
    top_vendedores,
    top_clientes,
    ventas_por_region,
    ventas_por_categoria,
    ventas_por_segmento,
)


# ── ESTILOS REUTILIZABLES ─────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT  = Font(color="1F4E78", bold=True, size=14)
BORDER_THIN = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
CENTER = Alignment(horizontal='center', vertical='center')


# ── FUNCIONES DE FORMATO ───────────────────────────────────────────────────────

def format_header(ws, row=1):
    """Aplica formato de encabezado a la fila indicada."""
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER_THIN


def autofit_columns(ws):
    """Ajusta automáticamente el ancho de las columnas según contenido."""
    for col_cells in ws.columns:
        length = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=10
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(length + 4, 40)


def apply_borders(ws, n_rows, n_cols):
    """Aplica bordes finos a un rango de datos."""
    for row in ws.iter_rows(min_row=1, max_row=n_rows, min_col=1, max_col=n_cols):
        for cell in row:
            cell.border = BORDER_THIN


def write_df(writer, df, sheet_name, start_row=0):
    """Escribe un DataFrame en una hoja y aplica formato básico."""
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
    ws = writer.sheets[sheet_name]
    format_header(ws, row=start_row + 1)
    apply_borders(ws, n_rows=len(df) + 1 + start_row, n_cols=len(df.columns))
    autofit_columns(ws)
    return ws


# ── HOJA: RESUMEN EJECUTIVO ────────────────────────────────────────────────────

def build_resumen_sheet(writer, df):
    """Crea la hoja de resumen ejecutivo con KPIs generales."""
    generales = kpis_generales(df)

    resumen_df = pd.DataFrame({
        'KPI': [
            'Ventas Totales',
            'Ganancia Total',
            'Costo Total',
            'Margen Promedio (%)',
            'Ticket Promedio',
            'Unidades Vendidas',
            'Total Transacciones',
            'Clientes Únicos',
            'Productos Vendidos',
        ],
        'Valor': [
            generales['VentasTotales'],
            generales['GananciaTotal'],
            generales['CostoTotal'],
            generales['MargenPromedio'],
            generales['TicketPromedio'],
            generales['UnidadesVendidas'],
            generales['TotalTransacciones'],
            generales['ClientesUnicos'],
            generales['ProductosVendidos'],
        ]
    })

    sheet_name = 'Resumen Ejecutivo'
    resumen_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
    ws = writer.sheets[sheet_name]

    # Título
    ws['A1'] = '📊 Resumen Ejecutivo - Sales Intelligence Platform'
    ws['A1'].font = TITLE_FONT

    # Fecha de generación
    ws['A2'] = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True, size=9, color="808080")

    format_header(ws, row=3)
    apply_borders(ws, n_rows=len(resumen_df) + 3, n_cols=2)
    autofit_columns(ws)

    # Formato de moneda para valores monetarios
    moneda_rows = [4, 5, 6, 8]  # Ventas, Ganancia, Costo, Ticket
    for r in moneda_rows:
        ws[f'B{r}'].number_format = '$#,##0.00'


# ── EJECUCIÓN PRINCIPAL ───────────────────────────────────────────────────────

if __name__ == "__main__":
    print("🚀 Generando reporte Excel...\n")

    df = load_dataset()
    os.makedirs(DATA_EXPORTS, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_path = f"{DATA_EXPORTS}/reporte_ejecutivo_{timestamp}.xlsx"

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        print("  → Hoja: Resumen Ejecutivo")
        build_resumen_sheet(writer, df)

        print("  → Hoja: Tendencia Mensual")
        write_df(writer, ventas_por_mes(df), 'Tendencia Mensual')

        print("  → Hoja: Comparativa Anual")
        write_df(writer, comparativa_anual(df), 'Comparativa Anual')

        print("  → Hoja: Top Productos")
        write_df(writer, top_productos(df, 10), 'Top Productos')

        print("  → Hoja: Top Vendedores")
        write_df(writer, top_vendedores(df, 10), 'Top Vendedores')

        print("  → Hoja: Top Clientes")
        write_df(writer, top_clientes(df, 10), 'Top Clientes')

        print("  → Hoja: Ventas por Región")
        write_df(writer, ventas_por_region(df), 'Ventas por Región')

        print("  → Hoja: Ventas por Categoría")
        write_df(writer, ventas_por_categoria(df), 'Ventas por Categoría')

        print("  → Hoja: Ventas por Segmento")
        write_df(writer, ventas_por_segmento(df), 'Ventas por Segmento')

    print(f"\n💾 Reporte guardado en: {output_path}")
    print("\n🎉 Reporte Excel generado exitosamente")